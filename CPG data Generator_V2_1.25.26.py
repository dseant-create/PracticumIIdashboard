"""
CPG GFR 1000-row synthetic dataset generator (2023–2027)
- Creates 1000 rows with the requested columns + calculations
- Writes an Excel file with light formatting (freeze header, table style, number formats)

Usage (VS Code):
  python generate_cpg_gfr_table.py

Output:
  CPG_GFR_Table_1000.xlsx
"""

import random
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

# Excel formatting (optional but included)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# -------------------------
# Configuration
# -------------------------
N_ROWS = 1000
OUTPUT_XLSX = "CPG_GFR_Table_1000.xlsx"

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

SUPPLIER_COUNTRIES = ["USA", "Mexico", "China", "India"]

SUPPLIERS = [
    ("Procter & Gamble", "USA"),
    ("Unilever", "USA"),
    ("Kimberly-Clark", "USA"),
    ("Nestlé", "USA"),
    ("PepsiCo", "USA"),
    ("The Coca-Cola Company", "USA"),
    ("Colgate-Palmolive", "USA"),
    ("Johnson & Johnson", "USA"),
    ("General Mills", "USA"),
    ("Kraft Heinz", "USA"),
    ("Mondelez International", "USA"),
    ("The Clorox Company", "USA"),
    ("Reckitt", "USA"),
    ("SC Johnson", "USA"),
    ("Church & Dwight", "USA"),
    ("Henkel", "USA"),
    ("Frito-Lay", "USA"),
    ("Conagra Brands", "USA"),
    ("Tyson Foods", "USA"),
    ("Ferrero", "USA"),
    ("Grupo Bimbo", "Mexico"),
    ("Gruma", "Mexico"),
    ("Maruchan", "China"),
    ("ITC Limited", "India"),
    ("Dabur", "India"),
    ("Godrej Consumer Products", "India"),
]

COMPONENTS_POOL = [
    "Raw Materials",
    "Packaging",
    "Labor",
    "Freight",
    "Overhead",
    "Ingredients",
    "Resin",
    "Pulp",
    "Active Chemicals",
    "Fragrance",
    "Co-manufacturing",
    "Utilities",
    "Quality Testing",
    "Warehousing",
]

# (Department Name Group, Product Type, Brand Options, Size Options)
PRODUCT_TEMPLATES = [
    ("Grocery", "Cereal", ["Cheerios", "Corn Flakes", "Frosted Flakes", "Honey Oats"], ["12 oz", "18 oz", "24 oz"]),
    ("Grocery", "Snack Bars", ["Nature Valley", "Kind", "Quaker", "Clif"], ["6 ct", "12 ct", "18 ct"]),
    ("Grocery", "Potato Chips", ["Lay's", "Ruffles", "Pringles", "Kettle"], ["5 oz", "8 oz", "12 oz"]),
    ("Grocery", "Pasta", ["Barilla", "Ronzoni", "De Cecco", "Store Brand"], ["12 oz", "16 oz"]),
    ("Grocery", "Pasta Sauce", ["Rao's", "Prego", "Classico", "Store Brand"], ["24 oz", "32 oz"]),
    ("Grocery", "Coffee", ["Folgers", "Dunkin'", "Starbucks", "Store Brand"], ["12 oz", "20 oz", "32 oz"]),
    ("Grocery", "Tea", ["Lipton", "Twinings", "Tazo", "Store Brand"], ["20 ct", "40 ct"]),
    ("Grocery", "Soda", ["Coca-Cola", "Pepsi", "Sprite", "Dr Pepper"], ["12 pk 12 oz", "6 pk 16.9 oz"]),
    ("Grocery", "Bottled Water", ["Aquafina", "Dasani", "Smartwater", "Store Brand"], ["24 pk", "12 pk"]),
    ("Grocery", "Yogurt", ["Chobani", "Dannon", "Oikos", "Store Brand"], ["4 pk", "6 oz"]),
    ("Household", "Paper Towels", ["Bounty", "Scott", "Viva", "Store Brand"], ["6 rolls", "8 rolls", "12 rolls"]),
    ("Household", "Toilet Paper", ["Charmin", "Cottonelle", "Scott", "Store Brand"], ["6 rolls", "12 rolls", "18 rolls"]),
    ("Household", "Trash Bags", ["Hefty", "Glad", "Store Brand", "Simplehuman"], ["20 ct", "40 ct", "60 ct"]),
    ("Household", "Laundry Detergent", ["Tide", "Gain", "Persil", "Store Brand"], ["92 oz", "150 oz", "64 loads"]),
    ("Household", "Dish Soap", ["Dawn", "Palmolive", "Seventh Generation", "Store Brand"], ["24 oz", "32 oz"]),
    ("Household", "Disinfecting Wipes", ["Clorox", "Lysol", "Store Brand"], ["75 ct", "110 ct"]),
    ("Personal Care", "Toothpaste", ["Crest", "Colgate", "Sensodyne", "Store Brand"], ["4.8 oz", "6 oz"]),
    ("Personal Care", "Shampoo", ["Head & Shoulders", "Pantene", "Dove", "Store Brand"], ["12 oz", "20 oz"]),
    ("Personal Care", "Body Wash", ["Dove", "Old Spice", "Olay", "Store Brand"], ["18 oz", "24 oz"]),
    ("Personal Care", "Deodorant", ["Old Spice", "Secret", "Dove", "Degree"], ["2.6 oz", "3 oz"]),
    ("Baby", "Diapers", ["Pampers", "Huggies", "Luvs", "Store Brand"], ["Size 3 72 ct", "Size 4 84 ct", "Size 5 64 ct"]),
    ("Baby", "Baby Wipes", ["Pampers", "Huggies", "WaterWipes", "Store Brand"], ["3 pk 168 ct", "1 pk 72 ct"]),
    ("Pet", "Dog Food", ["Purina", "Pedigree", "Blue Buffalo", "Store Brand"], ["15 lb", "30 lb", "40 lb"]),
    ("Pet", "Cat Litter", ["Tidy Cats", "Arm & Hammer", "Fresh Step", "Store Brand"], ["14 lb", "20 lb", "35 lb"]),
]


# -------------------------
# Helpers
# -------------------------
def random_upc(existing: set[str]) -> str:
    """Generate a unique 12-digit UPC."""
    while True:
        upc = str(random.randint(10**11, 10**12 - 1))
        if upc not in existing:
            existing.add(upc)
            return upc


def random_date_in_year(year: int) -> date:
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    return start + timedelta(days=random.randint(0, (end - start).days))


def base_price_for(product_type: str) -> float:
    """Heuristic base price by product type."""
    if product_type == "Diapers":
        return round(random.uniform(18.99, 44.99), 2)
    if product_type == "Dog Food":
        return round(random.uniform(14.99, 49.99), 2)
    if product_type == "Cat Litter":
        return round(random.uniform(9.99, 24.99), 2)
    if product_type == "Laundry Detergent":
        return round(random.uniform(8.99, 26.99), 2)
    if product_type in {"Paper Towels", "Toilet Paper", "Trash Bags", "Disinfecting Wipes"}:
        return round(random.uniform(5.49, 22.99), 2)
    if product_type in {"Soda", "Bottled Water"}:
        return round(random.uniform(3.99, 12.99), 2)
    return round(random.uniform(1.49, 14.99), 2)


def allocate_price_patterns(n: int) -> list[str]:
    """
    Price pattern rules:
      - flat: 15% keep same all years
      - up: 30% increase YoY
      - down: 20% decrease YoY
      - mixed: remainder varies
    """
    patterns = (["flat"] * int(0.15 * n)) + (["up"] * int(0.30 * n)) + (["down"] * int(0.20 * n))
    while len(patterns) < n:
        patterns.append("mixed")
    random.shuffle(patterns)
    return patterns


def generate_prices(p23: float, pattern: str) -> tuple[float, float, float, float, float]:
    if pattern == "flat":
        return p23, p23, p23, p23, p23

    if pattern == "up":
        r1, r2, r3, r4 = random.uniform(0.02, 0.12), random.uniform(0.02, 0.12), random.uniform(0.02, 0.12), random.uniform(0.02, 0.12)
        p24 = round(p23 * (1 + r1), 2)
        p25 = round(p24 * (1 + r2), 2)
        p26 = round(p25 * (1 + r3), 2)
        p27 = round(p26 * (1 + r4), 2)
        return p23, p24, p25, p26, p27

    if pattern == "down":
        r1, r2, r3, r4 = random.uniform(0.02, 0.12), random.uniform(0.02, 0.12), random.uniform(0.02, 0.12), random.uniform(0.02, 0.12)
        p24 = round(max(0.49, p23 * (1 - r1)), 2)
        p25 = round(max(0.49, p24 * (1 - r2)), 2)
        p26 = round(max(0.49, p25 * (1 - r3)), 2)
        p27 = round(max(0.49, p26 * (1 - r4)), 2)
        return p23, p24, p25, p26, p27

    # mixed (alternating up and down patterns)
    r1, r2, r3, r4 = random.uniform(0.02, 0.12), random.uniform(0.02, 0.12), random.uniform(0.02, 0.12), random.uniform(0.02, 0.12)
    if random.random() < 0.5:
        p24 = round(p23 * (1 + r1), 2)
        p25 = round(max(0.49, p24 * (1 - r2)), 2)
        p26 = round(p25 * (1 + r3), 2)
        p27 = round(max(0.49, p26 * (1 - r4)), 2)
    else:
        p24 = round(max(0.49, p23 * (1 - r1)), 2)
        p25 = round(p24 * (1 + r2), 2)
        p26 = round(max(0.49, p25 * (1 - r3)), 2)
        p27 = round(p26 * (1 + r4), 2)
    return p23, p24, p25, p26, p27


def clip_margin(x: float) -> float:
    return float(np.clip(x, 0.05, 0.50))


def write_excel_with_formatting(df: pd.DataFrame, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "CPG_GFR_1000"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws.freeze_panes = "A2"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Basic column widths (tweak as needed)
    widths = {
        "A": 12, "B": 12, "C": 14, "D": 36, "E": 22, "F": 12, "G": 16,
        "H": 9, "I": 9, "J": 12, "K": 14, "L": 14, "M": 14, "N": 14, "O": 14,
        "P": 14, "Q": 14, "R": 14, "S": 16, "T": 16, "U": 16, "V": 16,
        "W": 13, "X": 13, "Y": 13, "Z": 13, "AA": 13, "AB": 13, "AC": 18,
        "AD": 18, "AE": 14, "AF": 18, "AG": 14, "AH": 18, "AI": 14,
        "AJ": 15, "AK": 15, "AL": 15, "AM": 15, "AN": 15, "AO": 15,
        "AP": 15, "AQ": 15, "AR": 15, "AS": 15, "AT": 15, "AU": 15,
        "AV": 15, "AW": 15, "AX": 15, "AY": 15, "AZ": 15,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Number formats
    currency = '"$"#,##0.00'
    units_fmt = "#,##0"
    pct_fmt = "0.00%"
    cost_fmt = '"$"#,##0.0000'
    date_fmt = "yyyy-mm-dd"

    col_index = {name: idx + 1 for idx, name in enumerate(df.columns)}

    def set_col_format(col_name: str, fmt: str) -> None:
        col = col_index[col_name]
        for row in range(2, len(df) + 2):
            ws.cell(row=row, column=col).number_format = fmt

    for cname in [
        "Sale Price 2023", "Sale Price 2024", "Sale Price 2025", "Sale Price 2026", "Sale Price 2027",
        "Total Sales $ 2023", "Total Sales $ 2024", "Total Sales $ 2025", "Total Sales $ 2026", "Total Sales $ 2027",
        "Unit Margin $ 2023", "Unit Margin $ 2024", "Unit Margin $ 2025", "Unit Margin $ 2026", "Unit Margin $ 2027",
    ]:
        set_col_format(cname, currency)

    for cname in [
        "Unit Cost 2023", "Unit Cost 2024", "Unit Cost 2025", "Unit Cost 2026", "Unit Cost 2027",
        "Comp1 Cost 2023", "Comp1 Cost 2024", "Comp1 Cost 2025", "Comp1 Cost 2026", "Comp1 Cost 2027",
        "Comp2 Cost 2023", "Comp2 Cost 2024", "Comp2 Cost 2025", "Comp2 Cost 2026", "Comp2 Cost 2027",
        "Comp3 Cost 2023", "Comp3 Cost 2024", "Comp3 Cost 2025", "Comp3 Cost 2026", "Comp3 Cost 2027",
    ]:
        set_col_format(cname, cost_fmt)

    for cname in ["Sales Units 2023", "Sales Units 2024", "Sales Units 2025", "Sales Units 2026", "Sales Units 2027"]:
        set_col_format(cname, units_fmt)

    for cname in ["Margin % 2023", "Margin % 2024", "Margin % 2025", "Margin % 2026", "Margin % 2027", "Component 1 %", "Component 2 %", "Component 3 %"]:
        set_col_format(cname, pct_fmt)

    set_col_format("Last Cost Change Date", date_fmt)

    # Right-align numerics
    for c in range(1, len(df.columns) + 1):
        name = df.columns[c - 1]
        if any(k in name for k in ["Price", "Units", "Sales", "Margin", "Cost", "%", "#", "ID"]):
            for r in range(2, len(df) + 2):
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")

    # Add an Excel table
    table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    tab = Table(displayName="CPG_GFR_Table", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    ws.auto_filter.ref = table_ref

    wb.save(out_path)


# -------------------------
# Main generator
# -------------------------
def main() -> None:
    patterns = allocate_price_patterns(N_ROWS)
    upc_set: set[str] = set()
    rows: list[dict] = []

    for i in range(N_ROWS):
        # Required general fields
        item_category = "GFR"
        dept_num = random.randint(1, 20)

        store_num = random.randint(1, 50)
        region_num = ((store_num - 1) // 10) + 1  # 1..5
        store_country = "USA"

        # Product description
        _, product_type, brands, sizes = random.choice(PRODUCT_TEMPLATES)
        brand = random.choice(brands)
        size = random.choice(sizes)

        flavor = ""
        if product_type in {"Yogurt", "Snack Bars", "Cereal", "Tea"}:
            flavor = random.choice(["", " - Vanilla", " - Strawberry", " - Honey", " - Mixed Berry", " - Original"])

        item_desc = f"{brand} {product_type}{flavor}, {size}"
        upc = random_upc(upc_set)

        # Supplier fields
        supplier_name, default_country = random.choice(SUPPLIERS)
        supplier_id = random.randint(100000, 999999)
        supplier_country = default_country if default_country in SUPPLIER_COUNTRIES else random.choice(SUPPLIER_COUNTRIES)

        # Prices per year with required pattern mix
        p23 = base_price_for(product_type)
        p23, p24, p25, p26, p27 = generate_prices(p23, patterns[i])

        # Units per year
        u23 = random.randint(10_000, 100_000)
        u24 = int(round(u23 * random.uniform(0.80, 1.20)))
        u25 = int(round(u24 * random.uniform(0.80, 1.20)))
        u26 = int(round(u25 * random.uniform(0.80, 1.20)))
        u27 = int(round(u26 * random.uniform(0.80, 1.20)))
        u24 = min(100_000, max(10_000, u24))
        u25 = min(100_000, max(10_000, u25))
        u26 = min(100_000, max(10_000, u26))
        u27 = min(100_000, max(10_000, u27))

        # Margin % per year (5%–50%)
        m23 = clip_margin(np.random.normal(0.28, 0.10))
        m24 = clip_margin(m23 + np.random.normal(0.00, 0.05))
        m25 = clip_margin(m24 + np.random.normal(0.00, 0.05))
        m26 = clip_margin(m25 + np.random.normal(0.00, 0.05))
        m27 = clip_margin(m26 + np.random.normal(0.00, 0.05))
        m23, m24, m25, m26, m27 = round(m23, 4), round(m24, 4), round(m25, 4), round(m26, 4), round(m27, 4)

        # Unit costs derived from margin% and price: cost = price * (1 - margin%)
        c23 = round(p23 * (1 - m23), 4)
        c24 = round(p24 * (1 - m24), 4)
        c25 = round(p25 * (1 - m25), 4)
        c26 = round(p26 * (1 - m26), 4)
        c27 = round(p27 * (1 - m27), 4)

        # Unit margin $: price - cost
        mu23 = round(p23 - c23, 4)
        mu24 = round(p24 - c24, 4)
        mu25 = round(p25 - c25, 4)
        mu26 = round(p26 - c26, 4)
        mu27 = round(p27 - c27, 4)

        # Total sales $ = units * price
        s23 = round(u23 * p23, 2)
        s24 = round(u24 * p24, 2)
        s25 = round(u25 * p25, 2)
        s26 = round(u26 * p26, 2)
        s27 = round(u27 * p27, 2)

        # Last cost change date within last 5 years, aligned to year of change
        if c27 != c26:
            last_year = 2027
        elif c26 != c25:
            last_year = 2026
        elif c25 != c24:
            last_year = 2025
        elif c24 != c23:
            last_year = 2024
        else:
            last_year = 2023
        last_cost_change_date = random_date_in_year(last_year).isoformat()

        # 3-component breakdown: % sums to 1.0
        comps = random.sample(COMPONENTS_POOL, 3)
        perc = np.random.dirichlet([2.2, 2.0, 1.8])
        p1 = round(float(perc[0]), 4)
        p2 = round(float(perc[1]), 4)
        p3 = round(1.0 - (p1 + p2), 4)  # force exact sum after rounding

        # Component costs per year = unit cost * component %
        comp1_23, comp2_23, comp3_23 = round(c23 * p1, 4), round(c23 * p2, 4), round(c23 * p3, 4)
        comp1_24, comp2_24, comp3_24 = round(c24 * p1, 4), round(c24 * p2, 4), round(c24 * p3, 4)
        comp1_25, comp2_25, comp3_25 = round(c25 * p1, 4), round(c25 * p2, 4), round(c25 * p3, 4)
        comp1_26, comp2_26, comp3_26 = round(c26 * p1, 4), round(c26 * p2, 4), round(c26 * p3, 4)
        comp1_27, comp2_27, comp3_27 = round(c27 * p1, 4), round(c27 * p2, 4), round(c27 * p3, 4)

        rows.append(
            {
                "Item Category": item_category,
                "Department #": dept_num,
                "UPC": upc,
                "Item Description": item_desc,
                "Supplier Name": supplier_name,
                "Supplier ID": supplier_id,
                "Supplier Country": supplier_country,
                "Store #": store_num,
                "Region #": region_num,
                "Store Country": store_country,
                "Sale Price 2023": p23,
                "Sale Price 2024": p24,
                "Sale Price 2025": p25,
                "Sale Price 2026": p26,
                "Sale Price 2027": p27,
                "Sales Units 2023": u23,
                "Sales Units 2024": u24,
                "Sales Units 2025": u25,
                "Sales Units 2026": u26,
                "Sales Units 2027": u27,
                "Total Sales $ 2023": s23,
                "Total Sales $ 2024": s24,
                "Total Sales $ 2025": s25,
                "Total Sales $ 2026": s26,
                "Total Sales $ 2027": s27,
                "Margin % 2023": m23,
                "Margin % 2024": m24,
                "Margin % 2025": m25,
                "Margin % 2026": m26,
                "Margin % 2027": m27,
                "Unit Margin $ 2023": mu23,
                "Unit Margin $ 2024": mu24,
                "Unit Margin $ 2025": mu25,
                "Unit Margin $ 2026": mu26,
                "Unit Margin $ 2027": mu27,
                "Unit Cost 2023": c23,
                "Unit Cost 2024": c24,
                "Unit Cost 2025": c25,
                "Unit Cost 2026": c26,
                "Unit Cost 2027": c27,
                "Last Cost Change Date": last_cost_change_date,
                "Component 1": comps[0],
                "Component 1 %": p1,
                "Component 2": comps[1],
                "Component 2 %": p2,
                "Component 3": comps[2],
                "Component 3 %": p3,
                "Comp1 Cost 2023": comp1_23,
                "Comp1 Cost 2024": comp1_24,
                "Comp1 Cost 2025": comp1_25,
                "Comp1 Cost 2026": comp1_26,
                "Comp1 Cost 2027": comp1_27,
                "Comp2 Cost 2023": comp2_23,
                "Comp2 Cost 2024": comp2_24,
                "Comp2 Cost 2025": comp2_25,
                "Comp2 Cost 2026": comp2_26,
                "Comp2 Cost 2027": comp2_27,
                "Comp3 Cost 2023": comp3_23,
                "Comp3 Cost 2024": comp3_24,
                "Comp3 Cost 2025": comp3_25,
                "Comp3 Cost 2026": comp3_26,
                "Comp3 Cost 2027": comp3_27,
            }
        )

    df = pd.DataFrame(rows)

    out_path = Path(OUTPUT_XLSX).resolve()
    write_excel_with_formatting(df, out_path)

    print(f"Wrote {len(df):,} rows to: {out_path}")


if __name__ == "__main__":
    main()
